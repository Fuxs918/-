from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.uix.anchorlayout import AnchorLayout
from kivy.uix.textinput import TextInput
from kivy.uix.spinner import Spinner
from kivy.uix.popup import Popup
from kivy.uix.image import Image
from kivy.uix.checkbox import CheckBox
from kivy.graphics import Color, Line, Rectangle, RoundedRectangle
from kivy.clock import Clock, mainthread
from kivy.core.window import Window
from kivy.utils import platform
from kivy.core.text import LabelBase
from kivy.graphics.texture import Texture
from kivy.properties import BooleanProperty

import datetime
import os
from openpyxl import Workbook
from PIL import Image as PILImage
import io

try:
    from pyzbar.pyzbar import decode
    PYZBAR_AVAILABLE = True
except:
    PYZBAR_AVAILABLE = False

try:
    from plyer import vibrator, flashlight
except ImportError:
    vibrator = None
    flashlight = None

try:
    import cv2
    CV2_AVAILABLE = True
except ImportError:
    CV2_AVAILABLE = False

if platform == 'android':
    from jnius import autoclass
    Environment = autoclass('android.os.Environment')
    FONT_PATH = '/system/fonts/NotoSansCJK-Regular.ttc'
else:
    FONT_PATH = 'C:/Windows/Fonts/msyh.ttc'
    Window.size = (400, 750)

try:
    LabelBase.register(name='ChineseFont', fn_regular=FONT_PATH)
    FONT_NAME = 'ChineseFont'
except:
    FONT_NAME = 'Roboto'


class CameraPreview(FloatLayout):
    """摄像头预览组件，带瞄准线和实时扫码"""
    scanning = BooleanProperty()
    
    def __init__(self, on_barcode_detected=None, on_flash_toggle=None, **kwargs):
        super().__init__(**kwargs)
        self.on_barcode_detected = on_barcode_detected
        self.on_flash_toggle = on_flash_toggle
        self.capture = None
        self.scanning = False
        self.last_scan_time = 0
        self.scan_cooldown = 1.5  # 扫码冷却时间（秒）
        self._flash_on = False
        
        # 创建图像显示区域
        self.img_widget = Image(size_hint=(1, 1), pos_hint={'x': 0, 'y': 0})
        self.add_widget(self.img_widget)
        
        # 闪光灯按钮（左上角）
        self.flash_btn = Button(
            text="闪光灯",
            size_hint=(None, None),
            size=(60, 32),
            pos_hint={'x': 0.02, 'top': 0.95},
            background_normal='',
            background_color=(0.2, 0.2, 0.2, 0.7),
            font_name=FONT_NAME,
            font_size=11,
            color=(1, 1, 1, 1)
        )
        self.flash_btn.bind(on_press=self._toggle_flash)
        # 使用 index=0 确保按钮在最上层
        self.add_widget(self.flash_btn, index=0)
        
        # 绘制瞄准线
        with self.canvas.after:
            self._draw_guide()
        
        self.bind(size=self._update_guide, pos=self._update_guide)
        
        # 启动摄像头
        Clock.schedule_once(self._start_camera, 0.5)
    
    def _toggle_flash(self, instance):
        """切换闪光灯"""
        self._flash_on = not self._flash_on
        if self._flash_on:
            instance.background_color = (0.9, 0.7, 0.2, 0.9)  # 黄色表示开启
        else:
            instance.background_color = (0.2, 0.2, 0.2, 0.7)  # 灰色表示关闭
        
        if self.on_flash_toggle:
            self.on_flash_toggle(self._flash_on)
    
    def _draw_guide(self):
        """绘制扫码瞄准线"""
        self.canvas.after.clear()
        with self.canvas.after:
            # 半透明遮罩
            Color(0, 0, 0, 0.3)
            self.mask_rects = []

    def _update_guide(self, *args):
        """更新瞄准线位置和大小"""
        self.canvas.after.clear()
        w, h = self.size
        px, py = self.pos
        
        if w <= 0 or h <= 0:
            return
        
        with self.canvas.after:
            cx, cy = px + w/2, py + h/2
            rect_w = min(w * 0.75, 280)
            rect_h = min(h * 0.5, 100)
            
            # 扫描区域边框
            Color(0.2, 0.55, 0.9, 0.8)
            Line(rectangle=(cx - rect_w, cy - rect_h, rect_w * 2, rect_h * 2), width=2)
            
            # 四角标记
            corner_len = min(rect_w, rect_h) * 0.2
            Color(0.2, 0.55, 0.9, 1)
            
            corners = [
                (cx - rect_w, cy + rect_h, cx - rect_w + corner_len, cy + rect_h, cx - rect_w, cy + rect_h - corner_len),
                (cx + rect_w - corner_len, cy + rect_h, cx + rect_w, cy + rect_h, cx + rect_w, cy + rect_h - corner_len),
                (cx - rect_w, cy - rect_h, cx - rect_w + corner_len, cy - rect_h, cx - rect_w, cy - rect_h + corner_len),
                (cx + rect_w - corner_len, cy - rect_h, cx + rect_w, cy - rect_h, cx + rect_w, cy - rect_h + corner_len),
            ]
            
            for c in corners:
                Line(points=[c[0], c[1], c[2], c[3]], width=3)
                Line(points=[c[0], c[1], c[4], c[5]], width=3)
            
            # 中心横线（瞄准线）
            Color(0.2, 0.55, 0.9, 0.8)
            line_len = min(rect_w, rect_h) * 0.3
            Line(points=[cx - line_len, cy, cx + line_len, cy], width=2)
    
    def _start_camera(self, dt):
        """启动摄像头"""
        if not CV2_AVAILABLE:
            print("OpenCV not available")
            return
        
        try:
            self.capture = cv2.VideoCapture(0)
            if self.capture.isOpened():
                self.scanning = True
                Clock.schedule_interval(self._update_frame, 1.0 / 30.0)  # 30 FPS
                print("Camera started")
            else:
                print("Failed to open camera")
        except Exception as e:
            print(f"Camera error: {e}")
    
    def _update_frame(self, dt):
        """更新摄像头画面"""
        if self.capture is None or not self.capture.isOpened():
            return False
        
        ret, frame = self.capture.read()
        if not ret:
            return True
        
        # 水平翻转（镜像效果）
        frame = cv2.flip(frame, 1)
        
        # 转换为 RGB
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        
        # 实时扫码
        if PYZBAR_AVAILABLE and self.scanning:
            self._try_decode(frame_rgb)
        
        # 转换为 Kivy Texture
        h, w = frame_rgb.shape[:2]
        texture = Texture.create(size=(w, h), colorfmt='rgb')
        texture.blit_buffer(frame_rgb.tobytes(), colorfmt='rgb', bufferfmt='ubyte')
        texture.flip_vertical()
        self.img_widget.texture = texture
        
        return True
    
    def _try_decode(self, frame):
        """尝试解码条码"""
        if not PYZBAR_AVAILABLE:
            return
        
        current_time = datetime.datetime.now().timestamp()
        if current_time - self.last_scan_time < self.scan_cooldown:
            return
        
        try:
            # 转换为 PIL Image 用于 pyzbar
            pil_image = PILImage.fromarray(frame)
            barcodes = decode(pil_image)
            
            for barcode in barcodes:
                code = barcode.data.decode('utf-8')
                if code:
                    self.last_scan_time = current_time
                    if self.on_barcode_detected:
                        self.on_barcode_detected(code)
                    break
        except Exception as e:
            print(f"Decode error: {e}")
    
    def stop(self):
        """停止摄像头"""
        self.scanning = False
        Clock.unschedule(self._update_frame)
        if self.capture:
            self.capture.release()
            self.capture = None
    
    def on_touch_down(self, touch):
        """处理触摸事件，让子组件（如按钮）优先接收"""
        # 先让子组件处理触摸事件（如按钮点击）
        for child in self.children:
            if child.collide_point(*touch.pos):
                if child.on_touch_down(touch):
                    return True
        
        # 如果子组件没有处理，且点击在预览区域，切换扫描状态
        if self.collide_point(*touch.pos):
            self.scanning = not self.scanning
            return True
        return super().on_touch_down(touch)


class CardLayout(BoxLayout):
    def __init__(self, bg_color=(1, 1, 1, 1), **kwargs):
        super().__init__(**kwargs)
        self.bg_color = bg_color
        with self.canvas.before:
            Color(*bg_color)
            self.bg_rect = RoundedRectangle(pos=self.pos, size=self.size, radius=[12])
        self.bind(size=self._update_rect, pos=self._update_rect)
    
    def _update_rect(self, *args):
        self.bg_rect.pos = self.pos
        self.bg_rect.size = self.size


class BarcodeItem(BoxLayout):
    def __init__(self, index, item_data, on_delete, on_select, **kwargs):
        super().__init__(orientation='horizontal', size_hint_y=None, height=50, spacing=8, **kwargs)
        self.index = index
        self.item_data = item_data
        self.on_delete = on_delete
        self.on_select = on_select
        self.selected = False
        
        # 选中复选框
        self.check_btn = Button(
            text="□",
            size_hint=(None, 1),
            width=35,
            background_normal='',
            background_color=(0.95, 0.97, 0.98, 1),
            font_name=FONT_NAME,
            font_size=16,
            color=(0.4, 0.5, 0.6, 1)
        )
        self.check_btn.bind(on_press=self.toggle_select)
        
        # 序号
        num_label = Label(
            text=f"{index}.",
            size_hint=(None, 1),
            width=35,
            font_name=FONT_NAME,
            font_size=14,
            color=(0.3, 0.5, 0.8, 1),
            halign='left',
            valign='middle'
        )
        num_label.bind(size=num_label.setter('text_size'))
        
        # 条码信息
        info_layout = BoxLayout(orientation='vertical', size_hint=(1, 1))
        code_label = Label(
            text=item_data['code'][:35],
            font_name=FONT_NAME,
            font_size=13,
            color=(0.2, 0.3, 0.4, 1),
            halign='left',
            valign='bottom'
        )
        code_label.bind(size=code_label.setter('text_size'))
        time_label = Label(
            text=item_data['time'],
            font_name=FONT_NAME,
            font_size=11,
            color=(0.5, 0.55, 0.6, 1),
            halign='left',
            valign='top'
        )
        time_label.bind(size=time_label.setter('text_size'))
        info_layout.add_widget(code_label)
        info_layout.add_widget(time_label)
        
        # 删除按钮
        del_btn = Button(
            text="删除",
            size_hint=(None, 1),
            width=50,
            background_normal='',
            background_color=(0.9, 0.4, 0.4, 1),
            font_name=FONT_NAME,
            font_size=12,
            color=(1, 1, 1, 1)
        )
        del_btn.bind(on_press=lambda x: on_delete(index - 1))
        
        self.add_widget(self.check_btn)
        self.add_widget(num_label)
        self.add_widget(info_layout)
        self.add_widget(del_btn)
    
    def toggle_select(self, instance):
        self.selected = not self.selected
        if self.selected:
            self.check_btn.text = "☑"
            self.check_btn.color = (0.2, 0.55, 0.9, 1)
        else:
            self.check_btn.text = "□"
            self.check_btn.color = (0.4, 0.5, 0.6, 1)
        self.on_select(self.index - 1, self.selected)


class SmartScannerApp(App):
    def build(self):
        self.groups = {'默认组': []}
        self.current_group = '默认组'
        self.code_sets = {'默认组': set()}
        self.auto_seq = False
        self.auto_seq_count = 5  # 默认连号数量
        self.current_seq_index = 0
        self.selected_items = set()
        self.camera_preview = None

        root = BoxLayout(orientation='vertical')
        with root.canvas.before:
            Color(0.92, 0.95, 0.98, 1)
            self.root_bg = Rectangle(pos=root.pos, size=root.size)
        root.bind(size=self._update_root_bg, pos=self._update_root_bg)

        header = BoxLayout(size_hint=(1, None), height=55, padding=[12, 8])
        with header.canvas.before:
            Color(0.25, 0.55, 0.85, 1)
            self.header_bg = Rectangle(pos=header.pos, size=header.size)
        header.bind(size=lambda i, v: setattr(self.header_bg, 'size', v), pos=lambda i, v: setattr(self.header_bg, 'pos', v))
        
        menu_btn = Button(
            text="菜单",
            size_hint=(None, 1),
            width=60,
            background_normal='',
            background_color=(0.2, 0.45, 0.75, 1),
            font_name=FONT_NAME,
            font_size=14,
            color=(1, 1, 1, 1)
        )
        menu_btn.bind(on_press=self.show_menu)
        
        self.title_label = Label(
            text="智能扫码器", 
            font_size=20, 
            bold=True, 
            font_name=FONT_NAME,
            color=(1, 1, 1, 1)
        )
        
        group_btn = Button(
            text="默认组",
            size_hint=(None, 1),
            width=80,
            background_normal='',
            background_color=(0.3, 0.55, 0.85, 1),
            font_name=FONT_NAME,
            font_size=13,
            color=(1, 1, 1, 1)
        )
        group_btn.bind(on_press=self.show_group_selector)
        self.group_btn = group_btn
        
        header.add_widget(menu_btn)
        header.add_widget(self.title_label)
        header.add_widget(group_btn)
        root.add_widget(header)

        content = BoxLayout(orientation='vertical', padding=15, spacing=12)

        # 1. 摄像头预览区域（上）
        scan_card = CardLayout(size_hint=(1, None), height=200, bg_color=(0.1, 0.1, 0.1, 1))
        if CV2_AVAILABLE:
            self.camera_preview = CameraPreview(
                on_barcode_detected=self.on_barcode_detected,
                on_flash_toggle=self.on_flash_toggle,
                size_hint=(1, 1)
            )
            scan_card.add_widget(self.camera_preview)
        else:
            # 显示提示信息
            hint = Label(
                text="摄像头预览需要安装:\npip install opencv-python pillow",
                font_name=FONT_NAME,
                font_size=14,
                color=(0.7, 0.7, 0.7, 1),
                halign='center'
            )
            scan_card.add_widget(hint)
        content.add_widget(scan_card)

        # 2. 状态栏
        self.status = Label(
            text="实时扫码模式 - 对准条码自动识别", 
            size_hint=(1, None), 
            height=25,
            color=(0.3, 0.4, 0.5, 1),
            font_name=FONT_NAME,
            font_size=13
        )
        content.add_widget(self.status)

        # 3. 扫码记录（中）
        record_card = CardLayout(size_hint=(1, 1), bg_color=(1, 1, 1, 1))
        
        # 标题栏
        title_layout = BoxLayout(size_hint=(1, None), height=40, padding=[10, 5])
        self.count_label = Label(
            text="扫码记录 (0条)",
            size_hint=(1, 1),
            font_name=FONT_NAME,
            font_size=15,
            color=(0.3, 0.4, 0.5, 1),
            halign='left',
            valign='middle'
        )
        self.count_label.bind(size=self.count_label.setter('text_size'))
        
        # 批量操作按钮
        self.batch_del_btn = Button(
            text="删除",
            size_hint=(None, 1),
            width=55,
            background_normal='',
            background_down='',
            background_disabled_normal='',
            background_disabled_down='',
            background_color=(0.75, 0.75, 0.8, 1),
            font_name=FONT_NAME,
            font_size=13,
            color=(1, 1, 1, 1),
            disabled=True
        )
        self.batch_del_btn.bind(on_press=self.delete_selected)
        
        self.select_all_btn = Button(
            text="全选",
            size_hint=(None, 1),
            width=55,
            background_normal='',
            background_color=(0.5, 0.65, 0.85, 1),
            font_name=FONT_NAME,
            font_size=13,
            color=(1, 1, 1, 1)
        )
        self.select_all_btn.bind(on_press=self.select_all_items)
        
        title_layout.add_widget(self.count_label)
        title_layout.add_widget(self.batch_del_btn)
        title_layout.add_widget(self.select_all_btn)
        
        # 条码列表
        self.scroll = ScrollView(size_hint=(1, 1))
        self.barcode_list = BoxLayout(orientation='vertical', size_hint_y=None, spacing=2, padding=[8, 0])
        self.barcode_list.bind(minimum_height=self.barcode_list.setter('height'))

        self.barcode_list.bind(minimum_height=self.barcode_list.setter('height'))
        self.scroll.add_widget(self.barcode_list)
        
        record_inner = BoxLayout(orientation='vertical', padding=[5, 5, 5, 10], spacing=5)
        record_inner.add_widget(title_layout)
        record_inner.add_widget(self.scroll)
        record_card.add_widget(record_inner)
        content.add_widget(record_card)

        # 4. 操作按钮（下）- 只保留扫码按钮
        btn_layout = BoxLayout(size_hint=(1, None), height=70, spacing=15)
        
        # 扫码按钮
        scan_btn = Button(
            text="扫码", 
            font_size=22, 
            background_normal='',
            background_color=(0.2, 0.55, 0.9, 1),
            font_name=FONT_NAME,
            color=(1, 1, 1, 1),
            bold=True
        )
        scan_btn.bind(on_press=self.manual_scan)
        
        btn_layout.add_widget(scan_btn)
        content.add_widget(btn_layout)

        root.add_widget(content)
        self.refresh_list()
        return root

    def _update_root_bg(self, instance, value):
        self.root_bg.pos = instance.pos
        self.root_bg.size = instance.size

    @mainthread
    def on_barcode_detected(self, code):
        """检测到条码时的回调"""
        if not code:
            return
        
        # 检查是否重复
        if code in self.code_sets[self.current_group]:
            self.update_status(f"重复条码: {code}")
            # 重复提示音
            self._play_beep(800, 0.3)  # 高频长音表示重复
            return
        
        # 正常扫码提示音
        self._play_beep(1200, 0.1)  # 短促提示音
        
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.groups[self.current_group].insert(0, {'code': code, 'time': timestamp})
        self.code_sets[self.current_group].add(code)
        
        # 连号模式：自动添加后续连号
        if self.auto_seq and self.current_seq_index < self.auto_seq_count - 1:
            self._add_sequence_codes(code)
        
        if vibrator:
            try:
                vibrator.vibrate(0.1)
            except:
                pass
        
        self.refresh_list()
        self.update_status(f"已扫码: {code}")
    
    def _add_sequence_codes(self, first_code):
        """添加连号条码"""
        # 尝试从条码中提取数字部分
        import re
        match = re.search(r'(\d+)$', first_code)
        if match:
            num_part = match.group(1)
            prefix = first_code[:match.start()]
            start_num = int(num_part)
            num_len = len(num_part)
            
            for i in range(1, self.auto_seq_count):
                next_num = start_num + i
                next_code = f"{prefix}{next_num:0{num_len}d}"
                
                # 如果已存在则跳过
                if next_code in self.code_sets[self.current_group]:
                    continue
                
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.groups[self.current_group].insert(0, {'code': next_code, 'time': timestamp})
                self.code_sets[self.current_group].add(next_code)
            
            self.current_seq_index = self.auto_seq_count
            self.update_status(f"已添加连号: {first_code} ~ {prefix}{start_num + self.auto_seq_count - 1:0{num_len}d}")
    
    def _play_beep(self, frequency, duration):
        """播放提示音"""
        try:
            if platform == 'android':
                # Android 使用系统提示音
                from jnius import autoclass
                ToneGenerator = autoclass('android.media.ToneGenerator')
                AudioManager = autoclass('android.media.AudioManager')
                tone = ToneGenerator(AudioManager.STREAM_NOTIFICATION, 100)
                tone.startTone(ToneGenerator.TONE_PROP_BEEP, int(duration * 1000))
            else:
                # PC 使用 winsound
                import winsound
                winsound.Beep(frequency, int(duration * 1000))
        except:
            pass

    def manual_scan(self, instance):
        """手动扫码 - 从当前摄像头画面识别"""
        if self.camera_preview and self.camera_preview.capture:
            # 临时降低冷却时间，立即扫码
            self.camera_preview.last_scan_time = 0
            self.update_status("正在识别...")
            
            # 尝试立即解码一帧
            ret, frame = self.camera_preview.capture.read()
            if ret:
                frame_rgb = cv2.cvtColor(cv2.flip(frame, 1), cv2.COLOR_BGR2RGB)
                if PYZBAR_AVAILABLE:
                    try:
                        pil_image = PILImage.fromarray(frame_rgb)
                        barcodes = decode(pil_image)
                        if barcodes:
                            code = barcodes[0].data.decode('utf-8')
                            self.on_barcode_detected(code)
                        else:
                            self.update_status("未识别到条码，请对准后重试")
                    except Exception as e:
                        self.update_status(f"识别失败: {e}")
                else:
                    self.update_status("扫码功能需要安装: pip install pyzbar")
            else:
                self.update_status("摄像头未就绪")
        else:
            # 摄像头不可用，使用 plyer 扫码
            self.start_scan(instance)

    def on_item_select(self, index, selected):
        if selected:
            self.selected_items.add(index)
        else:
            self.selected_items.discard(index)
        self.update_batch_buttons()
    
    def update_batch_buttons(self):
        count = len(self.selected_items)
        if count > 0:
            self.batch_del_btn.text = f"删除({count})"
            self.batch_del_btn.background_color = (0.9, 0.35, 0.35, 1)
            self.batch_del_btn.disabled = False
        else:
            self.batch_del_btn.text = "删除"
            self.batch_del_btn.background_color = (0.75, 0.75, 0.8, 1)
            self.batch_del_btn.disabled = True
    
    def delete_selected(self, instance):
        if not self.selected_items:
            return
        data = self.groups[self.current_group]
        for idx in sorted(self.selected_items, reverse=True):
            if 0 <= idx < len(data):
                code = data[idx]['code']
                data.pop(idx)
                self.code_sets[self.current_group].discard(code)
        self.selected_items.clear()
        self.update_batch_buttons()
        self.refresh_list()
        self.update_status(f"已删除 {len(self.selected_items)} 条记录")
    
    def select_all_items(self, instance):
        data = self.groups[self.current_group]
        if not data:
            return
        
        all_selected = len(self.selected_items) == len(data)
        
        if all_selected:
            self.selected_items.clear()
            instance.text = "全选"
        else:
            self.selected_items = set(range(len(data)))
            instance.text = "取消"
        
        self.refresh_list()
        self.update_batch_buttons()
    
    def delete_single_item(self, index):
        data = self.groups[self.current_group]
        if 0 <= index < len(data):
            code = data[index]['code']
            data.pop(index)
            self.code_sets[self.current_group].discard(code)
            self.selected_items.discard(index)
            self.update_batch_buttons()
            self.refresh_list()
            self.update_status(f"已删除: {code}")

    def refresh_list(self):
        self.barcode_list.clear_widgets()
        data = self.groups[self.current_group]
        self.count_label.text = f"扫码记录 ({len(data)}条)"
        
        for i, item in enumerate(data):
            item_widget = BarcodeItem(
                index=i+1,
                item_data=item,
                on_delete=self.delete_single_item,
                on_select=self.on_item_select
            )
            if i in self.selected_items:
                item_widget.selected = True
                item_widget.check_btn.text = "☑"
                item_widget.check_btn.color = (0.2, 0.55, 0.9, 1)
            self.barcode_list.add_widget(item_widget)

    def update_status(self, text):
        self.status.text = text

    def on_flash_toggle(self, is_on):
        """处理闪光灯切换"""
        if flashlight:
            try:
                if is_on:
                    flashlight.on()
                    self.update_status("闪光灯已开启")
                else:
                    flashlight.off()
                    self.update_status("闪光灯已关闭")
            except Exception as e:
                self.update_status(f"闪光灯不支持: {e}")
        else:
            self.update_status("仅手机支持闪光灯")

    def show_menu(self, instance):
        content = BoxLayout(orientation='vertical', spacing=10, padding=15)
        
        # 1. 新建分组
        btn_new_group = Button(text="新建分组", font_name=FONT_NAME, background_color=(0.6, 0.5, 0.8, 1))
        btn_new_group.bind(on_press=lambda x: (self.create_new_group(), popup.dismiss()))
        
        # 2. 删除分组
        btn_delete_group = Button(text="删除分组", font_name=FONT_NAME, background_color=(0.9, 0.4, 0.4, 1))
        btn_delete_group.bind(on_press=lambda x: (self.show_delete_group_popup(), popup.dismiss()))
        
        # 3. 连号设置
        btn_auto = Button(text="连号设置", font_name=FONT_NAME, background_color=(0.5, 0.7, 0.5, 1))
        btn_auto.bind(on_press=lambda x: (self.show_auto_seq_popup(), popup.dismiss()))
        
        # 4. 导出Excel
        btn_export = Button(text="导出Excel", font_name=FONT_NAME, background_color=(0.3, 0.6, 0.9, 1))
        btn_export.bind(on_press=lambda x: (self.export_excel(), popup.dismiss()))
        
        content.add_widget(btn_new_group)
        content.add_widget(btn_delete_group)
        content.add_widget(btn_auto)
        content.add_widget(btn_export)
        
        # 添加标题标签（解决中文乱码）
        title_label = Label(
            text="菜单",
            font_name=FONT_NAME,
            size_hint=(1, None),
            height=30,
            color=(0.2, 0.4, 0.6, 1),
            bold=True
        )
        content.add_widget(title_label, index=len(content.children))
        
        popup = Popup(
            title="", 
            content=content, 
            size_hint=(0.85, 0.55),
            separator_color=(0.3, 0.5, 0.8, 1)
        )
        popup.open()

    def show_auto_seq_popup(self):
        content = BoxLayout(orientation='vertical', spacing=10, padding=15)
        
        # 添加标题标签（解决中文乱码）
        title_label = Label(
            text="连号设置",
            font_name=FONT_NAME,
            size_hint=(1, None),
            height=30,
            color=(0.2, 0.4, 0.6, 1),
            bold=True
        )
        content.add_widget(title_label)
        
        content.add_widget(Label(text="连号数量:", font_name=FONT_NAME, size_hint=(1, None), height=30))
        count_input = TextInput(text=str(self.auto_seq_count), multiline=False, input_filter='int', font_name=FONT_NAME)
        content.add_widget(count_input)
        
        info_label = Label(
            text="开启后，扫第一个条码后自动连扫指定数量",
            font_name=FONT_NAME,
            size_hint=(1, None),
            height=40,
            color=(0.5, 0.5, 0.5, 1)
        )
        content.add_widget(info_label)
        
        btn_toggle = Button(
            text="关闭连号" if self.auto_seq else "开启连号",
            font_name=FONT_NAME,
            background_color=(0.9, 0.5, 0.4, 1) if self.auto_seq else (0.4, 0.7, 0.5, 1)
        )
        
        def toggle_auto(instance):
            self.auto_seq_count = int(count_input.text or '5')
            self.auto_seq = not self.auto_seq
            self.current_seq_index = 0
            self.update_status(f"连号模式{'开启' if self.auto_seq else '关闭'}: 连扫{self.auto_seq_count}个")
            popup.dismiss()
        
        btn_toggle.bind(on_press=toggle_auto)
        content.add_widget(btn_toggle)
        
        popup = Popup(
            title="", 
            content=content, 
            size_hint=(0.8, 0.5),
            separator_color=(0.3, 0.5, 0.8, 1)
        )
        popup.open()

    def show_group_selector(self, instance):
        content = BoxLayout(orientation='vertical', spacing=8, padding=10)
        
        # 添加标题标签（解决中文乱码）
        title_label = Label(
            text="选择分组",
            font_name=FONT_NAME,
            size_hint=(1, None),
            height=30,
            color=(0.2, 0.4, 0.6, 1),
            bold=True
        )
        content.add_widget(title_label)
        
        for group_name in self.groups.keys():
            btn = Button(
                text=f"{'● ' if group_name == self.current_group else '  '}{group_name} ({len(self.groups[group_name])}条)",
                font_name=FONT_NAME,
                background_color=(0.3, 0.6, 0.9, 1) if group_name == self.current_group else (0.85, 0.9, 0.95, 1),
                color=(1, 1, 1, 1) if group_name == self.current_group else (0.3, 0.4, 0.5, 1)
            )
            btn.bind(on_press=lambda x, name=group_name: self.switch_group(name, popup))
            content.add_widget(btn)
        
        popup = Popup(
            title="", 
            content=content, 
            size_hint=(0.8, 0.65),
            separator_color=(0.3, 0.5, 0.8, 1)
        )
        popup.open()

    def switch_group(self, name, popup):
        self.current_group = name
        self.group_btn.text = name
        self.selected_items.clear()
        self.update_batch_buttons()
        self.refresh_list()
        popup.dismiss()
        self.update_status(f"切换到: {name}")

    def create_new_group(self):
        content = BoxLayout(orientation='vertical', spacing=10, padding=15)
        
        # 添加标题标签（解决中文乱码）
        title_label = Label(
            text="新建分组",
            font_name=FONT_NAME,
            size_hint=(1, None),
            height=30,
            color=(0.2, 0.4, 0.6, 1),
            bold=True
        )
        content.add_widget(title_label)
        
        text_input = TextInput(hint_text="分组名称", multiline=False, font_name=FONT_NAME)
        content.add_widget(text_input)
        
        def do_create(instance):
            name = text_input.text.strip()
            if name and name not in self.groups:
                self.groups[name] = []
                self.code_sets[name] = set()
                self.switch_group(name, popup)
            popup.dismiss()
        
        btn = Button(text="创建", font_name=FONT_NAME, background_color=(0.3, 0.6, 0.9, 1))
        btn.bind(on_press=do_create)
        content.add_widget(btn)
        
        popup = Popup(
            title="", 
            content=content, 
            size_hint=(0.8, 0.4),
            separator_color=(0.3, 0.5, 0.8, 1)
        )
        popup.open()

    def show_delete_group_popup(self):
        """显示删除分组弹窗"""
        content = BoxLayout(orientation='vertical', spacing=10, padding=15)
        
        # 添加标题标签（解决中文乱码）
        title_label = Label(
            text="删除分组",
            font_name=FONT_NAME,
            size_hint=(1, None),
            height=30,
            color=(0.2, 0.4, 0.6, 1),
            bold=True
        )
        content.add_widget(title_label)
        
        content.add_widget(Label(text="选择要删除的分组:", font_name=FONT_NAME, size_hint=(1, None), height=30))
        
        scroll = ScrollView(size_hint=(1, 1))
        groups_layout = BoxLayout(orientation='vertical', spacing=5, size_hint_y=None)
        groups_layout.bind(minimum_height=groups_layout.setter('height'))
        
        for group_name in list(self.groups.keys()):
            if group_name == "默认组":
                continue  # 默认组不能删除
            
            box = BoxLayout(size_hint_y=None, height=45)
            lbl = Label(
                text=f"{group_name} ({len(self.groups[group_name])}条)",
                font_name=FONT_NAME,
                halign='left',
                size_hint=(0.7, 1)
            )
            lbl.bind(size=lambda x, s: setattr(x, 'text_size', s))
            
            btn_delete = Button(
                text="删除",
                font_name=FONT_NAME,
                size_hint=(0.3, 1),
                background_color=(0.9, 0.4, 0.4, 1)
            )
            btn_delete.bind(on_press=lambda x, g=group_name: self._delete_group(g, popup))
            
            box.add_widget(lbl)
            box.add_widget(btn_delete)
            groups_layout.add_widget(box)
        
        scroll.add_widget(groups_layout)
        content.add_widget(scroll)
        
        btn_close = Button(text="关闭", font_name=FONT_NAME, size_hint=(1, None), height=45)
        btn_close.bind(on_press=lambda x: popup.dismiss())
        content.add_widget(btn_close)
        
        popup = Popup(
            title="",
            content=content,
            size_hint=(0.85, 0.65),
            separator_color=(0.3, 0.5, 0.8, 1)
        )
        popup.open()
    
    def _delete_group(self, group_name, popup):
        """删除指定分组"""
        if group_name in self.groups:
            del self.groups[group_name]
            del self.code_sets[group_name]
            
            # 如果删除的是当前组，切换到默认组
            if self.current_group == group_name:
                self.current_group = "默认组"
                self.refresh_list()
            
            self.update_status(f"已删除分组: {group_name}")
            popup.dismiss()
            self.show_delete_group_popup()  # 刷新列表

    def start_scan(self, instance):
        """使用 plyer 调用系统扫码"""
        try:
            from plyer import barcode
            barcode.scan(self._on_scan_result)
        except:
            self.update_status("扫码功能需要手机支持")

    def export_excel(self):
        """显示导出选项弹窗"""
        content = BoxLayout(orientation='vertical', spacing=10, padding=15)
        
        # 添加标题标签（解决中文乱码）
        title_label = Label(
            text="导出Excel",
            font_name=FONT_NAME,
            size_hint=(1, None),
            height=30,
            color=(0.2, 0.4, 0.6, 1),
            bold=True
        )
        content.add_widget(title_label)
        
        # 组选择区域
        content.add_widget(Label(text="选择要导出的组:", font_name=FONT_NAME, size_hint=(1, None), height=30))
        
        scroll = ScrollView(size_hint=(1, 1))
        groups_layout = BoxLayout(orientation='vertical', spacing=5, size_hint_y=None)
        groups_layout.bind(minimum_height=groups_layout.setter('height'))
        
        self.export_selected_groups = []
        checkboxes = []
        
        for group_name in self.groups.keys():
            box = BoxLayout(size_hint_y=None, height=40)
            chk = CheckBox(size_hint_x=None, width=40)
            chk.group_name = group_name
            chk.bind(active=lambda x, v, g=group_name: self._toggle_export_group(g, v))
            checkboxes = checkboxes
            count = len(self.groups[group_name])
            lbl = Label(text=f"{group_name} ({count}条)", font_name=FONT_NAME, halign='left')
            lbl.bind(size=lambda x, s: setattr(x, 'text_size', s))
            box.add_widget(chk)
            box.add_widget(lbl)
            groups_layout.add_widget(box)
        
        scroll.add_widget(groups_layout)
        content.add_widget(scroll)
        
        # 全选按钮
        btn_select_all = Button(text="全选", font_name=FONT_NAME, size_hint=(1, None), height=40, background_color=(0.5, 0.7, 0.5, 1))
        btn_select_all.bind(on_press=lambda x: self._select_all_groups(checkboxes, True))
        content.add_widget(btn_select_all)
        
        # 导出路径
        content.add_widget(Label(text="导出位置:", font_name=FONT_NAME, size_hint=(1, None), height=25))
        path_layout = BoxLayout(size_hint=(1, None), height=40, spacing=5)
        self.export_path_input = TextInput(
            text=os.path.expanduser("~"),
            multiline=False,
            font_name=FONT_NAME,
            size_hint=(0.75, 1)
        )
        btn_browse = Button(
            text="浏览...",
            font_name=FONT_NAME,
            size_hint=(0.25, 1),
            background_color=(0.5, 0.5, 0.6, 1)
        )
        btn_browse.bind(on_press=lambda x: self._browse_export_path())
        path_layout.add_widget(self.export_path_input)
        path_layout.add_widget(btn_browse)
        content.add_widget(path_layout)
        
        content.add_widget(Label(text="文件名:", font_name=FONT_NAME, size_hint=(1, None), height=25))
        default_name = f"扫码数据_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.export_filename_input = TextInput(text=default_name, multiline=False, font_name=FONT_NAME, size_hint=(1, None), height=40)
        content.add_widget(self.export_filename_input)
        
        # 按钮区域
        btn_layout = BoxLayout(size_hint=(1, None), height=50, spacing=10)
        btn_cancel = Button(text="取消", font_name=FONT_NAME)
        btn_cancel.bind(on_press=lambda x: popup.dismiss())
        btn_ok = Button(text="导出", font_name=FONT_NAME, background_color=(0.3, 0.6, 0.9, 1))
        btn_ok.bind(on_press=lambda x: self._do_export(popup))
        btn_layout.add_widget(btn_cancel)
        btn_layout.add_widget(btn_ok)
        content.add_widget(btn_layout)
        
        popup = Popup(
            title="",
            content=content,
            size_hint=(0.9, 0.75),
            separator_color=(0.3, 0.5, 0.8, 1)
        )
        popup.open()
    
    def _toggle_export_group(self, group_name, is_active):
        """切换组选择状态"""
        if is_active:
            if group_name not in self.export_selected_groups:
                self.export_selected_groups.append(group_name)
        else:
            if group_name in self.export_selected_groups:
                self.export_selected_groups.remove(group_name)
    
    def _select_all_groups(self, checkboxes, select):
        """全选/取消全选"""
        for chk in checkboxes:
            chk.active = select
    
    def _browse_export_path(self):
        """浏览选择导出路径"""
        try:
            if platform == 'android':
                # Android 使用 plyer 的文件选择器
                from plyer import filechooser
                filechooser.choose_dir(on_selection=self._on_path_selected)
            else:
                # PC 使用 tkinter
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk()
                root.withdraw()
                path = filedialog.askdirectory()
                root.destroy()
                if path:
                    self.export_path_input.text = path
        except Exception as e:
            self.update_status(f"选择路径失败: {e}")
    
    def _on_path_selected(self, selection):
        """Android 路径选择回调"""
        if selection:
            self.export_path_input.text = selection[0]
    
    def _do_export(self, popup):
        """执行导出"""
        if not self.export_selected_groups:
            self.update_status("请至少选择一个组")
            return
        
        filename = self.export_filename_input.text.strip()
        if not filename:
            filename = f"扫码数据_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        # 获取导出路径
        export_path = self.export_path_input.text.strip()
        if not export_path or not os.path.exists(export_path):
            if platform == 'android':
                export_path = os.path.join(Environment.getExternalStorageDirectory().getAbsolutePath(), 'Download')
            else:
                export_path = os.path.expanduser('~')
        
        try:
            wb = Workbook()
            for group_name in self.export_selected_groups:
                items = self.groups.get(group_name, [])
                ws = wb.create_sheet(title=group_name[:31])
                ws.append(["序号", "条码内容", "扫描时间"])
                for i, item in enumerate(items, 1):
                    ws.append([i, item['code'], item['time']])
            
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
                wb.remove(wb['Sheet'])
            
            full_path = os.path.join(export_path, filename)
            wb.save(full_path)
            popup.dismiss()
            self.update_status(f"导出成功: {full_path}")
        except Exception as e:
            self.update_status(f"导出失败: {e}")

    def on_stop(self):
        """应用关闭时释放摄像头"""
        if self.camera_preview:
            self.camera_preview.stop()


if __name__ == '__main__':
    SmartScannerApp().run()